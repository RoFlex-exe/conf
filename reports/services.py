# reports/services.py
import csv
import json
from datetime import datetime, timedelta
from io import BytesIO
from django.http import HttpResponse
from django.db.models import Count, Sum, Avg, Q
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
import pandas as pd

from conferences.models import Conference, ConferenceApplication
from organizations.models import Organization
from users.models import CustomUser


class ReportGenerator:
    """
    Класс для генерации различных отчётов
    """

    @staticmethod
    def generate_conferences_report(parameters, format='excel'):
        """
        Генерация отчёта по конференциям
        """
        # Базовый запрос
        queryset = Conference.objects.all().select_related('organization')

        # Применяем фильтры
        if parameters.get('date_from'):
            queryset = queryset.filter(start_date__gte=parameters['date_from'])
        if parameters.get('date_to'):
            queryset = queryset.filter(end_date__lte=parameters['date_to'])
        if parameters.get('status'):
            queryset = queryset.filter(status=parameters['status'])
        if parameters.get('organization'):
            queryset = queryset.filter(organization_id=parameters['organization'])
        if parameters.get('conference_type'):
            queryset = queryset.filter(conference_type=parameters['conference_type'])

        # Подготавливаем данные
        data = []
        for conf in queryset:
            data.append({
                'ID': conf.id,
                'Название': conf.title,
                'Краткое название': conf.short_title,
                'Организация': conf.organization.name,
                'Тип': conf.get_conference_type_display(),
                'Формат': conf.get_format_display(),
                'Дата начала': conf.start_date.strftime('%d.%m.%Y'),
                'Дата окончания': conf.end_date.strftime('%d.%m.%Y'),
                'Дедлайн': conf.deadline.strftime('%d.%m.%Y'),
                'Место': conf.location,
                'Статус': conf.get_status_display(),
                'Просмотры': conf.view_count,
                'В избранном': conf.favorites_count,
                'Заявок': conf.applications_count,
                'Рекомендуемая': 'Да' if conf.is_featured else 'Нет',
                'Создана': conf.created_at.strftime('%d.%m.%Y %H:%M'),
            })

        # Генерируем файл в нужном формате
        if format == 'excel':
            return ReportGenerator._to_excel(data, 'Конференции')
        elif format == 'csv':
            return ReportGenerator._to_csv(data, 'conferences')
        elif format == 'json':
            return ReportGenerator._to_json(data)
        elif format == 'pdf':
            return ReportGenerator._to_pdf(data, 'Отчёт по конференциям')

    @staticmethod
    def generate_applications_report(parameters, format='excel'):
        """
        Генерация отчёта по заявкам
        """
        queryset = ConferenceApplication.objects.all().select_related(
            'conference', 'user', 'conference__organization'
        )

        if parameters.get('date_from'):
            queryset = queryset.filter(created_at__date__gte=parameters['date_from'])
        if parameters.get('date_to'):
            queryset = queryset.filter(created_at__date__lte=parameters['date_to'])
        if parameters.get('status'):
            queryset = queryset.filter(status=parameters['status'])
        if parameters.get('conference'):
            queryset = queryset.filter(conference_id=parameters['conference'])
        if parameters.get('organization'):
            queryset = queryset.filter(conference__organization_id=parameters['organization'])

        data = []
        for app in queryset:
            data.append({
                'ID': app.id,
                'ФИО': app.full_name,
                'Email': app.email,
                'Организация участника': app.organization,
                'Конференция': app.conference.title,
                'Организатор': app.conference.organization.name,
                'Тема доклада': app.presentation_title,
                'Тип доклада': app.get_presentation_type_display(),
                'Статус': app.get_status_display(),
                'Дата подачи': app.created_at.strftime('%d.%m.%Y %H:%M'),
                'Комментарий': app.comment,
            })

        if format == 'excel':
            return ReportGenerator._to_excel(data, 'Заявки')
        elif format == 'csv':
            return ReportGenerator._to_csv(data, 'applications')
        elif format == 'json':
            return ReportGenerator._to_json(data)
        elif format == 'pdf':
            return ReportGenerator._to_pdf(data, 'Отчёт по заявкам')

    @staticmethod
    def generate_organizations_report(parameters, format='excel'):
        """
        Генерация отчёта по организациям
        """
        queryset = Organization.objects.all().annotate(
            conferences_count=Count('conferences'),
            total_views=Sum('conferences__view_count'),
            total_applications=Sum('conferences__applications_count')
        )

        if parameters.get('is_active') is not None:
            queryset = queryset.filter(is_active=parameters['is_active'])
        if parameters.get('is_verified') is not None:
            queryset = queryset.filter(is_verified=parameters['is_verified'])

        data = []
        for org in queryset:
            data.append({
                'ID': org.id,
                'Название': org.name,
                'Краткое название': org.short_name,
                'ИНН': org.inn,
                'Контактное лицо': org.contact_person,
                'Email': org.contact_email,
                'Телефон': org.contact_phone,
                'Сайт': org.website,
                'Активна': 'Да' if org.is_active else 'Нет',
                'Верифицирована': 'Да' if org.is_verified else 'Нет',
                'Конференций': org.conferences_count,
                'Всего просмотров': org.total_views or 0,
                'Всего заявок': org.total_applications or 0,
                'Дата регистрации': org.created_at.strftime('%d.%m.%Y'),
            })

        if format == 'excel':
            return ReportGenerator._to_excel(data, 'Организации')
        elif format == 'csv':
            return ReportGenerator._to_csv(data, 'organizations')
        elif format == 'json':
            return ReportGenerator._to_json(data)
        elif format == 'pdf':
            return ReportGenerator._to_pdf(data, 'Отчёт по организациям')

    @staticmethod
    def generate_users_report(parameters, format='excel'):
        """
        Генерация отчёта по пользователям
        """
        queryset = CustomUser.objects.all()

        if parameters.get('date_joined_from'):
            queryset = queryset.filter(date_joined__date__gte=parameters['date_joined_from'])
        if parameters.get('date_joined_to'):
            queryset = queryset.filter(date_joined__date__lte=parameters['date_joined_to'])
        if parameters.get('is_active') is not None:
            queryset = queryset.filter(is_active=parameters['is_active'])
        if parameters.get('email_verified') is not None:
            queryset = queryset.filter(email_verified=parameters['email_verified'])

        data = []
        for user in queryset:
            data.append({
                'ID': user.id,
                'Email': user.email,
                'Логин': user.username,
                'ФИО': user.get_full_name(),
                'Место работы': user.affiliation,
                'Статус': user.get_academic_degree_display() if user.academic_degree else '',
                'Email подтверждён': 'Да' if user.email_verified else 'Нет',
                'Активен': 'Да' if user.is_active else 'Нет',
                'Дата регистрации': user.date_joined.strftime('%d.%m.%Y'),
                'Последний вход': user.last_login.strftime('%d.%m.%Y %H:%M') if user.last_login else '',
            })

        if format == 'excel':
            return ReportGenerator._to_excel(data, 'Пользователи')
        elif format == 'csv':
            return ReportGenerator._to_csv(data, 'users')
        elif format == 'json':
            return ReportGenerator._to_json(data)
        elif format == 'pdf':
            return ReportGenerator._to_pdf(data, 'Отчёт по пользователям')

    @staticmethod
    def generate_statistics_report(parameters, format='excel'):
        """
        Генерация статистического отчёта
        """
        today = timezone.now().date()

        # Общая статистика
        total_conferences = Conference.objects.count()
        total_published = Conference.objects.filter(status='published').count()
        total_organizations = Organization.objects.filter(is_active=True).count()
        total_users = CustomUser.objects.count()
        total_applications = ConferenceApplication.objects.count()

        # Статистика по месяцам
        last_12_months = []
        for i in range(11, -1, -1):
            month_start = (today.replace(day=1) - timedelta(days=30 * i))
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

            conferences_created = Conference.objects.filter(
                created_at__date__gte=month_start,
                created_at__date__lte=month_end
            ).count()

            applications_created = ConferenceApplication.objects.filter(
                created_at__date__gte=month_start,
                created_at__date__lte=month_end
            ).count()

            last_12_months.append({
                'Месяц': month_start.strftime('%B %Y'),
                'Новых конференций': conferences_created,
                'Новых заявок': applications_created,
            })

        # Статистика по тематикам
        from conferences.models import Topic
        topic_stats = []
        for topic in Topic.objects.filter(is_active=True):
            conf_count = topic.conferences.count()
            if conf_count > 0:
                topic_stats.append({
                    'Тематика': topic.name,
                    'Конференций': conf_count,
                })

        # Статистика по организациям (топ-10)
        org_stats = []
        for org in Organization.objects.filter(is_active=True).annotate(
                conf_count=Count('conferences')
        ).order_by('-conf_count')[:10]:
            org_stats.append({
                'Организация': org.name,
                'Конференций': org.conf_count,
            })

        data = {
            'Общая статистика': [{
                'Показатель': 'Всего конференций',
                'Значение': total_conferences
            }, {
                'Показатель': 'Опубликовано',
                'Значение': total_published
            }, {
                'Показатель': 'Активных организаций',
                'Значение': total_organizations
            }, {
                'Показатель': 'Пользователей',
                'Значение': total_users
            }, {
                'Показатель': 'Заявок',
                'Значение': total_applications
            }],
            'По месяцам': last_12_months,
            'По тематикам': topic_stats,
            'Топ организаций': org_stats,
        }

        if format == 'excel':
            return ReportGenerator._to_excel_multisheet(data, 'Статистика')
        elif format == 'json':
            return ReportGenerator._to_json(data)
        elif format == 'pdf':
            return ReportGenerator._to_pdf_statistics(data)

    @staticmethod
    def _to_excel(data, sheet_name):
        """Генерация Excel файла"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        if not data:
            return wb

        # Заголовки
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Данные
        for row, item in enumerate(data, 2):
            for col, key in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=item[key])

        # Автоширина
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response[
            'Content-Disposition'] = f'attachment; filename="{sheet_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response

    @staticmethod
    def _to_excel_multisheet(data, filename):
        """Генерация Excel файла с несколькими листами"""
        wb = openpyxl.Workbook()

        for sheet_name, sheet_data in data.items():
            if sheet_name == list(data.keys())[0]:
                ws = wb.active
                ws.title = sheet_name[:31]
            else:
                ws = wb.create_sheet(title=sheet_name[:31])

            if not sheet_data:
                continue

            # Заголовки
            headers = list(sheet_data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

            # Данные
            for row, item in enumerate(sheet_data, 2):
                for col, key in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=item[key])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response[
            'Content-Disposition'] = f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response

    @staticmethod
    def _to_csv(data, filename):
        """Генерация CSV файла"""
        response = HttpResponse(content_type='text/csv')
        response[
            'Content-Disposition'] = f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        if not data:
            return response

        writer = csv.DictWriter(response, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

        return response

    @staticmethod
    def _to_json(data):
        """Генерация JSON файла"""
        response = HttpResponse(content_type='application/json')
        response[
            'Content-Disposition'] = f'attachment; filename="report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'

        json.dump(data, response, ensure_ascii=False, indent=2, default=str)

        return response

    @staticmethod
    def _to_pdf(data, title):
        """Генерация PDF файла"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []

        # Заголовок
        styles = getSampleStyleSheet()
        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 0.2 * inch))

        if not data:
            elements.append(Paragraph("Нет данных для отображения", styles['Normal']))
        else:
            # Подготавливаем данные для таблицы
            headers = list(data[0].keys())
            table_data = [headers]

            for item in data:
                row = [str(item.get(h, '')) for h in headers]
                table_data.append(row)

            # Создаём таблицу
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(table)

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type='application/pdf')
        response[
            'Content-Disposition'] = f'attachment; filename="report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        response.write(pdf)

        return response

    @staticmethod
    def _to_pdf_statistics(data):
        """Генерация PDF со статистикой"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []

        styles = getSampleStyleSheet()
        title_style = styles['Title']
        heading_style = styles['Heading2']
        normal_style = styles['Normal']

        # Заголовок
        elements.append(Paragraph("Статистический отчёт", title_style))
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(Paragraph(f"Сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}", normal_style))
        elements.append(Spacer(1, 0.3 * inch))

        # Общая статистика
        elements.append(Paragraph("Общая статистика", heading_style))
        elements.append(Spacer(1, 0.1 * inch))

        stats_data = [['Показатель', 'Значение']]
        for item in data['Общая статистика']:
            stats_data.append([item['Показатель'], str(item['Значение'])])

        stats_table = Table(stats_data)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 0.2 * inch))

        # По месяцам
        elements.append(Paragraph("Динамика по месяцам", heading_style))
        elements.append(Spacer(1, 0.1 * inch))

        monthly_data = [['Месяц', 'Новых конференций', 'Новых заявок']]
        for item in data['По месяцам']:
            monthly_data.append([
                item['Месяц'],
                str(item['Новых конференций']),
                str(item['Новых заявок'])
            ])

        monthly_table = Table(monthly_data)
        monthly_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(monthly_table)

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type='application/pdf')
        response[
            'Content-Disposition'] = f'attachment; filename="statistics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        response.write(pdf)

        return response